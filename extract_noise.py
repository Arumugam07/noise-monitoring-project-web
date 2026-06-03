#!/usr/bin/env python3
"""
Extract noise readings from Supabase: 2 Jun 2025 → 2 Jun 2026
Queries meter_readings directly, pivots in Python, exports to Excel.
"""

import os
import pandas as pd
from supabase import create_client
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell import WriteOnlyCell

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_ANON_KEY")
OUTPUT_FILE  = "noise_2025_2026.xlsx"
START_DT     = "2025-06-02T00:00:00"
END_DT       = "2026-06-02T23:59:59"
BATCH_SIZE   = 10000

LOCATION_NAMES = {
    "15490": "Singapore Sports School",
    "16034": "BLK 120 Serangoon North Ave 1",
    "16041": "BLK 838 Hougang Central",
    "14542": "BLK 558 Jurong West Street 42",
    "15725": "Jurong Safra Block C",
    "16032": "AMA KENG SITE",
    "16045": "BLK 19 Balam Road",
    "15820": "Norcom II Tower 4",
    "15821": "Blk 444 Choa Chu Kang Ave 4",
    "15999": "BLK 654B Punggol Drive",
    "16026": "BLK 132B Tengah Garden Avenue",
    "16004": "BLK 206A Punggol Place",
    "16005": "Woodlands 11",
}
LOCATION_IDS = list(LOCATION_NAMES.keys())


def fetch_all():
    client = create_client(SUPABASE_URL, SUPABASE_KEY)
    all_rows = []
    offset = 0

    print(f"Fetching from meter_readings: {START_DT} → {END_DT}")

    while True:
        resp = (
            client.table("meter_readings")
            .select("reading_datetime,location_id,reading_value")
            .gte("reading_datetime", START_DT)
            .lte("reading_datetime", END_DT)
            .in_("location_id", LOCATION_IDS)
            .order("reading_datetime", desc=False)
            .range(offset, offset + BATCH_SIZE - 1)
            .execute()
        )

        batch = resp.data or []
        if not batch:
            print(f"\nNo more data at offset {offset}")
            break

        all_rows.extend(batch)
        print(f"  {len(all_rows):,} rows...", end="\r")

        if len(batch) < BATCH_SIZE:
            break
        offset += BATCH_SIZE

    print(f"\nTotal rows: {len(all_rows):,}")

    # Debug: show first row to confirm column names
    if all_rows:
        print(f"Sample row keys: {list(all_rows[0].keys())}")
        print(f"Sample row: {all_rows[0]}")

    return all_rows


def build_pivot(rows):
    if not rows:
        raise ValueError("No data returned. Check date range and Supabase credentials.")

    print("Building pivot...")
    df = pd.DataFrame(rows)
    print(f"Columns in dataframe: {list(df.columns)}")

    # Find the datetime column (handle any naming)
    dt_col = None
    for col in df.columns:
        if "datetime" in col.lower() or "time" in col.lower():
            dt_col = col
            break
    if dt_col is None:
        raise ValueError(f"No datetime column found. Columns: {list(df.columns)}")

    print(f"Using datetime column: '{dt_col}'")

    # Convert to Singapore time
    df[dt_col] = pd.to_datetime(df[dt_col], utc=True)
    df[dt_col] = df[dt_col].dt.tz_convert("Asia/Singapore")

    df["Date"]   = df[dt_col].dt.strftime("%-d %b %y")
    df["Time"]   = df[dt_col].dt.strftime("%H:%M")
    df["dt_key"] = df[dt_col].dt.floor("min")

    df["reading_value"] = pd.to_numeric(df["reading_value"], errors="coerce")

    pivot = df.pivot_table(
        index=["dt_key", "Date", "Time"],
        columns="location_id",
        values="reading_value",
        aggfunc="max"
    ).reset_index()

    pivot = pivot.sort_values("dt_key").drop(columns=["dt_key"])
    pivot.columns.name = None

    rename = {lid: LOCATION_NAMES[lid] for lid in LOCATION_IDS if lid in pivot.columns}
    pivot = pivot.rename(columns=rename)

    loc_cols = [LOCATION_NAMES[lid] for lid in LOCATION_IDS if LOCATION_NAMES[lid] in pivot.columns]
    pivot = pivot[["Date", "Time"] + loc_cols]

    print(f"Pivot: {len(pivot):,} rows × {len(pivot.columns)} columns")
    return pivot


def db_color(val):
    if pd.isna(val): return None
    if val < 50:     return "D4EDDA"
    elif val < 70:   return "FFF3CD"
    elif val < 85:   return "FFE5D0"
    else:            return "F8D7DA"


def write_excel(df):
    print(f"Writing {OUTPUT_FILE}...")
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Noise Readings")

    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", start_color="1F77B4")
    date_fill = PatternFill("solid", start_color="EBF3FB")
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, bottom=thin)
    data_font = Font(name="Arial", size=9)
    center    = Alignment(horizontal="center", vertical="center")
    left      = Alignment(horizontal="left",   vertical="center")

    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 7
    for ci in range(3, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 24

    # Header row
    hdr_cells = []
    for h in df.columns:
        c = WriteOnlyCell(ws, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center; c.border = border
        hdr_cells.append(c)
    ws.append(hdr_cells)

    # Data rows
    for i, row in enumerate(df.itertuples(index=False), 1):
        cells = []
        for ci, val in enumerate(row):
            v = None if (isinstance(val, float) and pd.isna(val)) else val
            c = WriteOnlyCell(ws, value=v)
            c.font = data_font; c.border = border
            if ci == 0:
                c.fill = date_fill; c.alignment = left
            elif ci == 1:
                c.alignment = center
            else:
                c.alignment = center
                color = db_color(val)
                if color:
                    c.fill = PatternFill("solid", start_color=color)
            cells.append(c)
        ws.append(cells)

        if i % 100000 == 0:
            print(f"  Written {i:,} rows...")

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Saved: {OUTPUT_FILE}  ({len(df):,} rows)")


if __name__ == "__main__":
    rows  = fetch_all()
    pivot = build_pivot(rows)
    write_excel(pivot)
